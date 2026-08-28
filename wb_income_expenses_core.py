#!/usr/bin/env python3
"""
Парсинг, загрузка и сверка отчёта WB «Доходы и расходы».
Используется веб-формой (webapp/app.py).
"""

import re
from datetime import date
from pathlib import Path
from typing import Callable

import openpyxl
from dotenv import load_dotenv

from wb_core import get_client, SCRIPT_DIR  # get_client не дублируем

load_dotenv(SCRIPT_DIR / ".env")

# Маппинг заголовков xlsx → имена полей таблицы.
# Только «текущий период» — колонки с «(предыдущий период)» игнорируются.
_COL_MAP = {
    "Итог, ₽":                                          "total_rub",
    "Продажи, ₽":                                       "sales_rub",
    "Продажи, шт":                                      "n_sales",
    "Возвраты, ₽":                                      "returns_rub",
    "Возвраты, шт":                                     "n_returns",
    "Логистика, ₽":                                     "logistics_rub",
    "Штрафы, ₽":                                        "fines_rub",
    "Комиссия WB, ₽":                                   "commission_rub",
    "Эквайринг, ₽":                                     "acquiring_rub",
    "Потери, подмены и товары с дефектами, ₽":         "losses_rub",
    "Доплаты, ₽":                                       "bonuses_rub",
    "Программа лояльности, ₽":                         "loyalty_rub",
}

_INT_FIELDS = {"n_sales", "n_returns"}

INSERT_COLUMNS = [
    "cabinet", "period_start", "period_end",
    "n_sales", "n_returns",
    "sales_rub", "returns_rub", "logistics_rub", "fines_rub",
    "commission_rub", "acquiring_rub", "losses_rub", "bonuses_rub",
    "loyalty_rub", "total_rub",
    "source_file",
]

METRICS = [
    ("01 Кол-во продаж",  "kol_prodazh",  lambda r: r["n_sales"] - r["n_returns"],                5),
    ("02 Продажи + СПП",  "prodazhi_spp", lambda r: r["sales_rub"] + r["returns_rub"],           500),
    ("05 Комиссия ВБ",    "komissiya",    lambda r: r["commission_rub"] + r["acquiring_rub"],     500),
    ("07 Логистика",      "logistika",    lambda r: r["logistics_rub"],                           500),
    ("08 Штрафы",         "shtrafy",      lambda r: r["fines_rub"],                               100),
    ("09 Доплаты",        "doplaty",      lambda r: r["bonuses_rub"],                             100),
    ("13 Скидка Wibes",   "skidka_wibes", lambda r: r["loyalty_rub"],                             500),
    ("15 К перечислению", "k_perech",     lambda r: r["total_rub"],                              1000),
]


def parse_income_expenses(path: Path) -> dict:
    """
    Разбирает один xlsx «Доходы и расходы».
    Возвращает dict с period_start, period_end, source_file и суммами по колонкам.
    Raises ValueError если не удалось прочитать период или листы не найдены.
    """
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        raise ValueError(f"Не удалось открыть файл «{path.name}»: {e}") from e

    try:
        ws_info = wb["Общая информация"]
    except KeyError:
        raise ValueError(f"Лист «Общая информация» не найден в «{path.name}»")

    period_start = period_end = None
    for row in ws_info.iter_rows(values_only=True):
        if row and row[0] == "Выбранный период":
            m = re.match(r"С (\d{4}-\d{2}-\d{2}) по (\d{4}-\d{2}-\d{2})", str(row[1] or ""))
            if m:
                period_start = date.fromisoformat(m.group(1))
                period_end = date.fromisoformat(m.group(2))
            break

    if not period_start or not period_end:
        raise ValueError(f"Не удалось прочитать период из «{path.name}»")

    try:
        ws_data = wb["Детальная информация"]
    except KeyError:
        raise ValueError(f"Лист «Детальная информация» не найден в «{path.name}»")

    rows = list(ws_data.iter_rows(values_only=True))
    if len(rows) < 3:
        raise ValueError(f"Слишком мало строк в «{path.name}»")

    headers = rows[1]  # строка 2 — заголовки
    col_idx = {}
    for i, h in enumerate(headers):
        if h in _COL_MAP:
            col_idx[_COL_MAP[h]] = i

    totals: dict = {field: 0.0 for field in _COL_MAP.values()}
    for row in rows[2:]:
        for field, idx in col_idx.items():
            val = row[idx]
            if isinstance(val, (int, float)):
                totals[field] += float(val)

    for field in _INT_FIELDS:
        totals[field] = int(round(totals[field]))

    return {
        "period_start": period_start,
        "period_end": period_end,
        "source_file": path.name,
        **totals,
    }


def ingest_files(paths: list[Path], cabinet: str, log: Callable = print) -> dict:
    """
    Парсит список xlsx «Доходы и расходы» и загружает в wb_income_expenses.
    Повторная загрузка того же (cabinet, period_start) перезапишет запись.
    """
    rows_data = []
    for path in paths:
        parsed = parse_income_expenses(path)
        log(f"  {path.name}: {parsed['period_start']} — {parsed['period_end']}")
        row = [cabinet] + [parsed[col] for col in INSERT_COLUMNS[1:]]
        rows_data.append(row)

    if not rows_data:
        log("Нет файлов для загрузки.")
        return {"files": 0, "rows": 0}

    client = get_client()
    client.insert("wb_income_expenses", rows_data, column_names=INSERT_COLUMNS)
    log(f"Загружено {len(rows_data)} записей в wb_income_expenses.")

    return {"files": len(paths), "rows": len(rows_data)}


# SQL считает те же метрики, что wb_metrics_by_month.sql, но за произвольный диапазон дат.
# CS_K_TYPES — константа кода, не пользовательский ввод, поэтому зашита прямо в SQL.
_RECONCILE_SQL = """
WITH cs_k AS (SELECT arrayJoin([
    'продажа', 'сторно продаж', 'авансовая оплата за товар без движения',
    'возврат', 'корректный возврат', 'корректная продажа',
    'компенсация брака', 'компенсация потерянного товара',
    'сторно возвратов', 'компенсация ущерба',
    'добровольная компенсация при возврате',
    'компенсация подмененного товара', 'частичная компенсация брака'
]) AS v)
SELECT
    coalesce(sumIf(qty, lowerUTF8(trim(payment_reason)) = 'продажа'), 0)
    - coalesce(sumIf(qty, lowerUTF8(trim(payment_reason)) = 'возврат'), 0)
        AS kol_prodazh,

    coalesce(sumIf(retail_price_with_discount,
        lowerUTF8(trim(payment_reason)) IN (SELECT v FROM cs_k)
        AND lowerUTF8(trim(document_type)) = 'продажа'), 0)
    - coalesce(sumIf(retail_price_with_discount,
        lowerUTF8(trim(payment_reason)) IN (SELECT v FROM cs_k)
        AND lowerUTF8(trim(document_type)) = 'возврат'), 0)
        AS prodazhi_spp,

    (
      coalesce(sumIf(payable_to_seller,
          lowerUTF8(trim(payment_reason)) IN (SELECT v FROM cs_k)
          AND lowerUTF8(trim(document_type)) = 'продажа'), 0)
      - coalesce(sumIf(payable_to_seller,
          lowerUTF8(trim(payment_reason)) IN (SELECT v FROM cs_k)
          AND lowerUTF8(trim(document_type)) = 'возврат'), 0)
    ) - (
      coalesce(sumIf(retail_price_with_discount,
          lowerUTF8(trim(payment_reason)) IN (SELECT v FROM cs_k)
          AND lowerUTF8(trim(document_type)) = 'продажа'), 0)
      - coalesce(sumIf(retail_price_with_discount,
          lowerUTF8(trim(payment_reason)) IN (SELECT v FROM cs_k)
          AND lowerUTF8(trim(document_type)) = 'возврат'), 0)
    )
        AS komissiya,

    -(coalesce(sumIf(delivery_service_cost,
        payment_reason IN ('Логистика', 'Коррекция логистики')), 0))
        AS logistika,

    -(coalesce(sum(total_fines), 0)) AS shtrafy,

    -(coalesce(sum(wb_commission_correction), 0)) AS doplaty,

    coalesce(sum(loyalty_discount_compensation), 0)
    - coalesce(sum(loyalty_program_cost), 0)
    - coalesce(sum(loyalty_points_deducted), 0)
        AS skidka_wibes,

    (
      coalesce(sumIf(payable_to_seller,
          lowerUTF8(trim(payment_reason)) IN (SELECT v FROM cs_k)
          AND lowerUTF8(trim(document_type)) = 'продажа'), 0)
      - coalesce(sumIf(payable_to_seller,
          lowerUTF8(trim(payment_reason)) IN (SELECT v FROM cs_k)
          AND lowerUTF8(trim(document_type)) = 'возврат'), 0)
    )
    -(coalesce(sumIf(delivery_service_cost,
        payment_reason IN ('Логистика', 'Коррекция логистики')), 0))
    -(coalesce(sum(total_fines), 0))
    -(coalesce(sum(wb_commission_correction), 0))
    -(coalesce(sum(storage_cost), 0))
    -(coalesce(sum(acceptance_operations), 0))
    -(coalesce(sumIf(deductions,
        trim(REGEXP_REPLACE(REGEXP_REPLACE(logistics_fines_corrections_type,
            ',\\s*документ\\s*№\\s*\\d+', ''), '\\s+\\d+$', ''))
        NOT IN ('Оказание услуг «WB Продвижение»', 'Оказание услуг «ВБ.Продвижение»')
        OR logistics_fines_corrections_type IS NULL), 0))
    +(coalesce(sum(loyalty_discount_compensation), 0)
      - coalesce(sum(loyalty_program_cost), 0)
      - coalesce(sum(loyalty_points_deducted), 0))
    -(coalesce(sumIf(deductions,
        trim(REGEXP_REPLACE(REGEXP_REPLACE(logistics_fines_corrections_type,
            ',\\s*документ\\s*№\\s*\\d+', ''), '\\s+\\d+$', ''))
        IN ('Оказание услуг «WB Продвижение»', 'Оказание услуг «ВБ.Продвижение»')), 0))
        AS k_perech

FROM wb_reports FINAL
WHERE cabinet = {cabinet:String}
  AND sale_date IS NOT NULL
  AND sale_date >= {period_start:Date}
  AND sale_date <= {period_end:Date}
"""

_CH_ALIASES = ["kol_prodazh", "prodazhi_spp", "komissiya", "logistika",
               "shtrafy", "doplaty", "skidka_wibes", "k_perech"]


def reconcile_income_expenses(client, cabinet: str, log: Callable = print) -> list[tuple]:
    """
    Для каждого загруженного периода кабинета сравнивает 8 метрик
    из wb_income_expenses с расчётами по wb_reports.

    Возвращает list[tuple]:
    (period_start, period_end, metric_name, file_value, ch_value, diff, tolerance, is_ok)
    """
    ie_rows = client.query(
        "SELECT period_start, period_end, "
        "n_sales, n_returns, sales_rub, returns_rub, logistics_rub, fines_rub, "
        "commission_rub, acquiring_rub, losses_rub, bonuses_rub, loyalty_rub, total_rub "
        "FROM wb_income_expenses FINAL "
        "WHERE cabinet = {cabinet:String} ORDER BY period_start",
        parameters={"cabinet": cabinet},
    ).result_rows

    if not ie_rows:
        log(f"Нет записей в wb_income_expenses для cabinet='{cabinet}'")
        return []

    log(f"Сверка: {len(ie_rows)} период(ов) для кабинета '{cabinet}'")

    ie_fields = [
        "period_start", "period_end",
        "n_sales", "n_returns", "sales_rub", "returns_rub", "logistics_rub", "fines_rub",
        "commission_rub", "acquiring_rub", "losses_rub", "bonuses_rub", "loyalty_rub", "total_rub",
    ]

    results = []
    for ie_row in ie_rows:
        ie = dict(zip(ie_fields, ie_row))
        p_start = ie["period_start"]
        p_end = ie["period_end"]
        log(f"  {p_start} — {p_end}")

        ch_row = client.query(
            _RECONCILE_SQL,
            parameters={
                "cabinet": cabinet,
                "period_start": p_start,
                "period_end": p_end,
            },
        ).result_rows

        ch_values = dict(zip(_CH_ALIASES, ch_row[0])) if ch_row else {k: 0.0 for k in _CH_ALIASES}

        for metric_name, ch_alias, file_formula, tolerance in METRICS:
            file_val = float(file_formula(ie))
            ch_val = float(ch_values.get(ch_alias, 0.0))
            diff = abs(file_val - ch_val)
            is_ok = 1 if diff <= tolerance else 0
            label = "OK" if is_ok else f"MISMATCH {diff:.0f}"
            log(f"    [{label:>18}] {metric_name}  файл={file_val:.0f}  CH={ch_val:.0f}")
            results.append((p_start, p_end, metric_name, file_val, ch_val, diff, tolerance, is_ok))

    return results
