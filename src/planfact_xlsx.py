#!/usr/bin/env python3
"""
Парсер сырой выгрузки ПланФакта (xlsx, "Все операции" из planfact.io).

Формат: строка 1 — заголовок сервиса ("ПланФакт", ссылка), строка 2 — имена
колонок (20 колонок, A-T), с строки 3 — данные. Читаем по именам колонок,
а не по позиции — порядок колонок в экспорте ПланФакта не гарантирован.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl

SCRIPT_DIR = Path(__file__).parent

# имя колонки в файле -> имя поля в planfact_transactions
COLUMN_MAP = {
    "Дата оплаты": "payment_date",
    "Статус оплаты": "payment_status",
    "Дата начисления": "accrual_date",
    "Статус начисления": "accrual_status",
    "Контрагент": "counterparty",
    "ИНН контрагента": "counterparty_inn",
    "Тип": "operation_type",
    "Счет": "account_name",
    "№ Счета": "account_number",
    "Банк": "bank_name",
    "Бик": "bik",
    "Юрлицо": "legal_entity",
    "ИНН юрлица": "legal_entity_inn",
    "Статья": "statya",
    "Родительские статьи": "parent_statya",
    "Вид деятельности": "activity_type",
    "Назначение платежа": "payment_purpose",
    "Проекты": "pf_project",
    "Сумма": "amount",
    "Валюта": "currency",
}

DATE_FIELDS = {"payment_date", "accrual_date"}


def parse_xlsx(filepath: Path) -> list[dict]:
    """Разбирает один xlsx-экспорт ПланФакта и возвращает список строк-транзакций."""
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]

    rows_iter = ws.iter_rows(values_only=True)
    next(rows_iter)  # строка 1: заголовок сервиса
    header = next(rows_iter)  # строка 2: имена колонок

    field_by_index = {i: COLUMN_MAP[name] for i, name in enumerate(header) if name in COLUMN_MAP}

    rows = []
    for row_num, values in enumerate(rows_iter, start=1):
        if values[0] is None and all(v is None for v in values):
            continue
        row = {"row_num": row_num, "source_file": filepath.name}
        for idx, field in field_by_index.items():
            value = values[idx] if idx < len(values) else None
            if field in DATE_FIELDS and value is not None:
                value = value.date()
            row[field] = value
        rows.append(row)

    return rows


if __name__ == "__main__":
    import sys

    if len(sys.argv) < 2:
        print("Usage: planfact_xlsx.py <file.xlsx>", file=sys.stderr)
        sys.exit(1)

    all_rows = parse_xlsx(Path(sys.argv[1]))
    print(f"Строк: {len(all_rows)}")
    with_project = sum(1 for r in all_rows if r.get("pf_project"))
    print(f"Строк с заполненным pf_project: {with_project} ({with_project / len(all_rows):.0%})")
