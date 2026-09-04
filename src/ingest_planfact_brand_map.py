#!/usr/bin/env python3
"""
Загрузка справочников из гугл-таблицы Ильяса в planfact_brand_map
("Бренды") и planfact_accounts ("Счета"). Публичная таблица, без
API-ключа — качаем целиком как xlsx (не через gviz CSV: у него баг —
при экспорте листа "Счета" gviz схлопывает заголовок с первыми строками
данных, из-за смешанных типов колонок).

Пример:
    python3 ingest_planfact_brand_map.py --project-id 1
"""

import argparse
import io
import os
from pathlib import Path

import clickhouse_connect
import openpyxl
import requests
from dotenv import load_dotenv

SCRIPT_DIR = Path(__file__).parent
load_dotenv(SCRIPT_DIR.parent / ".env")

SHEET_ID = "1sgddmomCgTbSpcLItlW3t68h0ParmYXLU197nEzNVFA"
XLSX_URL = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"

# "Название в ПланФакте" в таблице "Бренды" разошлось с реальным текстом
# в выгрузке ПланФакта для Torado/Wb (переименование/опечатка, подтверждено
# Ильясом 2026-09-04) — здесь синоним нормализуется к варианту из выгрузки.
PF_PROJECT_ALIASES = {
    "WB Chromium (ИП Маторин)": "WB Torado (ИП Маторин)",
}

BRAND_MAP_COLUMNS = ["project_id", "pf_project", "brand", "platform", "legal_entity"]
ACCOUNTS_COLUMNS = ["project_id", "pf_account_name", "account_number", "account_label", "account_type"]


def get_client():
    host = os.environ["CLICKHOUSE_HOST"]
    port = int(os.environ.get("CLICKHOUSE_PORT", "8443"))
    user = os.environ.get("CLICKHOUSE_USER", "default")
    password = os.environ["CLICKHOUSE_PASSWORD"]
    database = os.environ.get("CLICKHOUSE_DATABASE", "default")
    secure = os.environ.get("CLICKHOUSE_SECURE", "1") != "0"
    return clickhouse_connect.get_client(
        host=host, port=port, username=user, password=password,
        database=database, secure=secure,
    )


def fetch_workbook() -> openpyxl.Workbook:
    resp = requests.get(XLSX_URL, timeout=30)
    resp.raise_for_status()
    return openpyxl.load_workbook(io.BytesIO(resp.content), data_only=True)


def to_text(value) -> str | None:
    """Числовые ячейки (например номер счёта, если Google Sheets сохранил его как число)
    приводим к целой строке, а не '4.08e+19'/'123.0' — иначе теряем/искажаем номер."""
    if value is None:
        return None
    if isinstance(value, float):
        return str(int(value))
    return str(value).strip() or None


def parse_brand_map(wb: openpyxl.Workbook) -> list[dict]:
    ws = wb["Бренды"]
    rows = []
    for brand, platform, pf_project, legal_entity in ws.iter_rows(min_row=2, values_only=True):
        pf_project = to_text(pf_project)
        brand = to_text(brand)
        if not pf_project or not brand:
            continue  # бренды без привязанного юрлица/проекта — нет ключа для джойна
        pf_project = PF_PROJECT_ALIASES.get(pf_project, pf_project)
        rows.append({
            "pf_project": pf_project,
            "brand": brand,
            "platform": to_text(platform) or "",
            "legal_entity": to_text(legal_entity),
        })
    return rows


def parse_accounts(wb: openpyxl.Workbook) -> list[dict]:
    ws = wb["Счета"]
    rows = []
    for pf_account_name, account_number, account_label, account_type in ws.iter_rows(min_row=2, values_only=True):
        pf_account_name = to_text(pf_account_name)
        if not pf_account_name:
            continue
        rows.append({
            "pf_account_name": pf_account_name,
            "account_number": to_text(account_number),
            "account_label": to_text(account_label),
            "account_type": to_text(account_type),
        })
    return rows


def main():
    parser = argparse.ArgumentParser(description="Загрузка справочников Бренды/Счета из ПланФакта")
    parser.add_argument("--project-id", required=True, type=int, help="ID проекта, например 1 для CloudSix")
    parser.add_argument("--dry-run", action="store_true", help="Не писать в ClickHouse, только проверить")
    args = parser.parse_args()

    wb = fetch_workbook()
    brand_rows = parse_brand_map(wb)
    account_rows = parse_accounts(wb)
    print(f"Бренды: {len(brand_rows)} строк, Счета: {len(account_rows)} строк")

    for row in brand_rows:
        row["project_id"] = args.project_id
    for row in account_rows:
        row["project_id"] = args.project_id

    if args.dry_run:
        for row in brand_rows:
            print("brand:", row)
        for row in account_rows:
            print("account:", row)
        print("Dry-run: в ClickHouse ничего не пишу.")
        return

    client = get_client()
    client.insert(
        "planfact_brand_map",
        [[row.get(c) for c in BRAND_MAP_COLUMNS] for row in brand_rows],
        column_names=BRAND_MAP_COLUMNS,
    )
    print(f"Загружено {len(brand_rows)} строк в planfact_brand_map.")

    client.insert(
        "planfact_accounts",
        [[row.get(c) for c in ACCOUNTS_COLUMNS] for row in account_rows],
        column_names=ACCOUNTS_COLUMNS,
    )
    print(f"Загружено {len(account_rows)} строк в planfact_accounts.")


if __name__ == "__main__":
    main()
